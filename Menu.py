import PyQt5.QtWidgets as qt
from PyQt5 import QtCore
import openpyxl
from game import Main


class MainWindow(qt.QWidget):
    '''glowne okno menu'''
    def __init__(self,app):
        super().__init__()
        self.setFixedSize(600,700)
        self.layout = qt.QVBoxLayout()
        self.button_start = qt.QPushButton("Start")
        self.layout.addWidget(self.button_start)
        self.button_start.clicked.connect(self.play_game)
        self.button_start.setFixedSize(570, 50)
        self.button_settings = qt.QPushButton("Settings")
        self.layout.addWidget(self.button_settings)
        self.button_settings.clicked.connect(self.on_show_settings)
        self.button_settings.setFixedSize(570, 50)
        self.button_scoreboard = qt.QPushButton("Scoreboard")
        self.layout.addWidget(self.button_scoreboard)
        self.button_scoreboard.clicked.connect(self.on_show_scoreboard)
        self.button_scoreboard.setFixedSize(570, 50)
        self.button_exit = qt.QPushButton("Exit")
        self.layout.addWidget(self.button_exit)
        self.button_exit.clicked.connect(qt.QApplication.instance().quit)
        self.button_exit.setFixedSize(570, 50)
        self.setLayout(self.layout)

    def on_show_scoreboard(self):
        '''przechodznie do innego okna z wynikami'''
        menu.close()
        scoreboard.show()

    def on_show_settings(self):
        '''przechodznie do innego okna z ustawieniami'''
        menu.close()
        settings.show()

    def play_game(self):
        '''przechodznie do innego okna z gra'''
        menu.close()
        game = Main(settings.mode)
        game.run()


class Scoreboard(qt.QWidget):
    '''okno z wynikami'''
    def __init__(self):
        super().__init__()
        self.setFixedSize(600, 700)
        self.layout = qt.QVBoxLayout()
        self.label = qt.QLabel("Score board:")
        self.label.move(100, 130)
        self.label.setAlignment(QtCore.Qt.AlignCenter)
        self.layout.addWidget(self.label)
        self.table_widget = qt.QTableWidget()
        self.layout.addWidget(self.table_widget)
        self.load_data()
        self.button2 = qt.QPushButton("MENU")
        self.button2.clicked.connect(self.on_show_menu)
        self.layout.addWidget(self.button2)
        self.setLayout(self.layout)

    def on_show_menu(self):
        '''przechodznie do innego okna z menu glowny'''
        scoreboard.close()
        menu.show()

    def load_data(self):
        '''wczytywanie z pliku informacji'''
        path = "score.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        self.table_widget.setRowCount(sheet.max_row)
        self.table_widget.setColumnCount(sheet.max_column)
        list_values = list(sheet.values)
        self.table_widget.setHorizontalHeaderLabels(list_values[0])
        row_index = 0
        for value_tuple in list_values[1:]:
            col_index = 0
            for value in value_tuple:
                self.table_widget.setItem(row_index, col_index, qt.QTableWidgetItem(str(value)))
                col_index += 1
            row_index += 1


class Settings(qt.QWidget):
    '''okno z ustawieniami'''
    def __init__(self):
        super().__init__()
        self.mode = 0
        self.setFixedSize(600, 700)
        self.layout = qt.QVBoxLayout()
        self.label = qt.QLabel("POZIOM TRUDNOSCI")
        self.label.setFixedSize(570, 50)
        self.label.move(100, 130)
        self.label.setAlignment(QtCore.Qt.AlignCenter)
        self.label2 = qt.QLabel("Easy")
        self.label2.move(100, 130)
        self.label2.setAlignment(QtCore.Qt.AlignCenter)
        self.label2.setFixedSize(570, 50)
        self.button_easy = qt.QPushButton("Easy")
        self.button_easy.clicked.connect(self.mode_easy)
        self.button_easy.setFixedSize(570, 50)
        self.button_medium = qt.QPushButton("Medium")
        self.button_medium.clicked.connect(self.mode_medium)
        self.button_medium.setFixedSize(570, 50)
        self.button_hard = qt.QPushButton("Hard")
        self.button_hard.clicked.connect(self.mode_hard)
        self.button_hard.setFixedSize(570, 50)
        self.button2 = qt.QPushButton("MENU")
        self.button2.clicked.connect(self.on_show_menu)
        self.button2.setFixedSize(570, 50)
        self.layout.addWidget(self.label)
        self.layout.addWidget(self.label2)
        self.layout.addWidget(self.button_easy)
        self.layout.addWidget(self.button_medium)
        self.layout.addWidget(self.button_hard)
        self.layout.addWidget(self.button2)
        self.setLayout(self.layout)

    def on_show_menu(self):
        '''przechodznie do innego okna z menu glowny'''
        settings.close()
        menu.show()

    def mode_easy(self):
        self.mode = 0
        self.label2.setText("Easy")

    def mode_medium(self):
        self.mode = 1
        self.label2.setText("Medium")

    def mode_hard(self):
        self.mode = 2
        self.label2.setText("Hard")


app = qt.QApplication([])
scoreboard = Scoreboard()
settings = Settings()
menu = MainWindow(app)
menu.show()
app.exec_()
